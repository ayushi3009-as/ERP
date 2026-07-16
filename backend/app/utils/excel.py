import io
from datetime import date, datetime
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
    numbers,
)
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="1A1A2E")
SUBTITLE_FONT = Font(name="Calibri", size=10, color="555555")
DATA_FONT = Font(name="Calibri", size=10)
BOLD_FONT = Font(name="Calibri", size=10, bold=True)
TOTAL_FILL = PatternFill(start_color="E8E8F0", end_color="E8E8F0", fill_type="solid")
ALT_FILL = PatternFill(start_color="F8F8FC", end_color="F8F8FC", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT_ALIGN = Alignment(horizontal="right", vertical="center")


def _num(val: Any) -> float:
    if val is None:
        return 0.0
    return float(val)


def _fmt_date(d: Any) -> str:
    if d is None:
        return ""
    if isinstance(d, str):
        return d
    if isinstance(d, (date, datetime)):
        return d.strftime("%d-%m-%Y")
    return str(d)


def _write_title(ws, row: int, title: str, subtitle: str = "") -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = TITLE_FONT
    cell.alignment = LEFT_ALIGN
    row += 1
    if subtitle:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        cell = ws.cell(row=row, column=1, value=subtitle)
        cell.font = SUBTITLE_FONT
        cell.alignment = LEFT_ALIGN
        row += 1
    return row + 1


def _write_headers(ws, row: int, headers: list[str], col_widths: list[int]) -> int:
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    return row + 1


def _write_data_row(
    ws, row: int, data: list, number_cols: Optional[set[int]] = None, alt: bool = False
) -> int:
    number_cols = number_cols or set()
    for col_idx, value in enumerate(data, 1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        cell.font = DATA_FONT
        cell.border = THIN_BORDER
        if col_idx in number_cols:
            cell.alignment = RIGHT_ALIGN
            cell.number_format = "#,##0.00"
        else:
            cell.alignment = LEFT_ALIGN
        if alt:
            cell.fill = ALT_FILL
    return row + 1


def _write_total_row(
    ws, row: int, data: list, number_cols: Optional[set[int]] = None
) -> int:
    number_cols = number_cols or set()
    for col_idx, value in enumerate(data, 1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        cell.font = BOLD_FONT
        cell.fill = TOTAL_FILL
        cell.border = THIN_BORDER
        if col_idx in number_cols:
            cell.alignment = RIGHT_ALIGN
            cell.number_format = "#,##0.00"
        else:
            cell.alignment = LEFT_ALIGN
    return row + 1


def _auto_filter(ws, start_row: int, end_row: int, num_cols: int):
    ws.auto_filter.ref = f"A{start_row}:{get_column_letter(num_cols)}{end_row}"


def _save_workbook(wb: Workbook) -> bytes:
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def generate_sales_report(
    records: list[dict],
    company_name: str = "",
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Report"

    row = 1
    subtitle_parts = []
    if from_date:
        subtitle_parts.append(f"From: {from_date}")
    if to_date:
        subtitle_parts.append(f"To: {to_date}")
    subtitle = " | ".join(subtitle_parts) if subtitle_parts else ""
    row = _write_title(ws, row, f"{company_name} - Sales Report", subtitle)

    headers = [
        "Date",
        "Invoice #",
        "Customer",
        "Items",
        "Subtotal",
        "Tax",
        "Discount",
        "Grand Total",
        "Paid",
        "Balance",
        "Status",
    ]
    widths = [12, 14, 22, 8, 14, 12, 12, 14, 14, 14, 12]
    row = _write_headers(ws, row, headers, widths)
    data_start = row

    total_subtotal = 0.0
    total_tax = 0.0
    total_discount = 0.0
    total_grand = 0.0
    total_paid = 0.0

    for idx, rec in enumerate(records):
        subtotal = _num(rec.get("subtotal"))
        tax = (
            _num(rec.get("cgst_amount"))
            + _num(rec.get("sgst_amount"))
            + _num(rec.get("igst_amount"))
        )
        discount = _num(rec.get("discount"))
        grand = _num(rec.get("grand_total"))
        paid = _num(rec.get("paid_amount"))

        total_subtotal += subtotal
        total_tax += tax
        total_discount += discount
        total_grand += grand
        total_paid += paid

        data_row = [
            _fmt_date(rec.get("invoice_date") or rec.get("order_date")),
            rec.get("invoice_number") or rec.get("order_number", ""),
            rec.get("customer_name", ""),
            rec.get("item_count", 0),
            subtotal,
            tax,
            discount,
            grand,
            paid,
            grand - paid,
            rec.get("status", ""),
        ]
        row = _write_data_row(
            ws, row, data_row, number_cols={5, 6, 7, 8, 9, 10}, alt=idx % 2 == 1
        )

    _auto_filter(ws, data_start, row - 1, len(headers))

    _write_total_row(
        ws,
        row,
        [
            "",
            "",
            "",
            "TOTAL",
            total_subtotal,
            total_tax,
            total_discount,
            total_grand,
            total_paid,
            total_grand - total_paid,
            "",
        ],
        number_cols={5, 6, 7, 8, 9, 10},
    )

    return _save_workbook(wb)


def generate_purchase_report(
    records: list[dict],
    company_name: str = "",
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Purchase Report"

    row = 1
    subtitle_parts = []
    if from_date:
        subtitle_parts.append(f"From: {from_date}")
    if to_date:
        subtitle_parts.append(f"To: {to_date}")
    subtitle = " | ".join(subtitle_parts) if subtitle_parts else ""
    row = _write_title(ws, row, f"{company_name} - Purchase Report", subtitle)

    headers = [
        "Date",
        "PO #",
        "Vendor",
        "Items",
        "Subtotal",
        "Tax",
        "Discount",
        "Grand Total",
        "Advance",
        "Balance",
        "Status",
    ]
    widths = [12, 14, 22, 8, 14, 12, 12, 14, 14, 14, 12]
    row = _write_headers(ws, row, headers, widths)
    data_start = row

    total_subtotal = 0.0
    total_tax = 0.0
    total_discount = 0.0
    total_grand = 0.0
    total_advance = 0.0

    for idx, rec in enumerate(records):
        subtotal = _num(rec.get("subtotal"))
        tax = (
            _num(rec.get("cgst_amount"))
            + _num(rec.get("sgst_amount"))
            + _num(rec.get("igst_amount"))
        )
        discount = _num(rec.get("discount"))
        grand = _num(rec.get("grand_total"))
        advance = _num(rec.get("advance_amount"))

        total_subtotal += subtotal
        total_tax += tax
        total_discount += discount
        total_grand += grand
        total_advance += advance

        data_row = [
            _fmt_date(rec.get("po_date")),
            rec.get("po_number", ""),
            rec.get("vendor_name", ""),
            rec.get("item_count", 0),
            subtotal,
            tax,
            discount,
            grand,
            advance,
            grand - advance,
            rec.get("status", ""),
        ]
        row = _write_data_row(
            ws, row, data_row, number_cols={5, 6, 7, 8, 9, 10}, alt=idx % 2 == 1
        )

    _auto_filter(ws, data_start, row - 1, len(headers))

    _write_total_row(
        ws,
        row,
        [
            "",
            "",
            "",
            "TOTAL",
            total_subtotal,
            total_tax,
            total_discount,
            total_grand,
            total_advance,
            total_grand - total_advance,
            "",
        ],
        number_cols={5, 6, 7, 8, 9, 10},
    )

    return _save_workbook(wb)


def generate_inventory_report(
    records: list[dict],
    company_name: str = "",
    warehouse_name: str = "",
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory Report"

    row = 1
    subtitle = f"Warehouse: {warehouse_name}" if warehouse_name else ""
    row = _write_title(ws, row, f"{company_name} - Inventory Report", subtitle)

    headers = [
        "SKU",
        "Product",
        "Category",
        "Unit",
        "Qty",
        "Reserved",
        "Damaged",
        "Available",
        "Avg Cost",
        "Total Value",
        "Reorder Lvl",
        "Status",
    ]
    widths = [12, 24, 14, 8, 10, 10, 10, 10, 12, 14, 12, 12]
    row = _write_headers(ws, row, headers, widths)
    data_start = row

    total_value = 0.0

    for idx, rec in enumerate(records):
        qty = _num(rec.get("quantity"))
        reserved = _num(rec.get("reserved_quantity"))
        damaged = _num(rec.get("damaged_quantity"))
        available = qty - reserved - damaged
        avg_cost = _num(rec.get("avg_cost"))
        value = available * avg_cost
        total_value += value

        reorder = _num(rec.get("reorder_level"))
        status = "Low Stock" if available <= reorder else "OK"

        data_row = [
            rec.get("sku", ""),
            rec.get("product_name", ""),
            rec.get("category_name", ""),
            rec.get("unit_code", ""),
            qty,
            reserved,
            damaged,
            available,
            avg_cost,
            value,
            reorder,
            status,
        ]
        row = _write_data_row(
            ws, row, data_row, number_cols={5, 6, 7, 8, 9, 10, 11}, alt=idx % 2 == 1
        )

    _auto_filter(ws, data_start, row - 1, len(headers))

    _write_total_row(
        ws,
        row,
        [
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            total_value,
            "",
            "",
        ],
        number_cols={10},
    )

    return _save_workbook(wb)


def generate_production_report(
    records: list[dict],
    company_name: str = "",
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Production Report"

    row = 1
    subtitle_parts = []
    if from_date:
        subtitle_parts.append(f"From: {from_date}")
    if to_date:
        subtitle_parts.append(f"To: {to_date}")
    subtitle = " | ".join(subtitle_parts) if subtitle_parts else ""
    row = _write_title(ws, row, f"{company_name} - Production Report", subtitle)

    headers = [
        "PO #",
        "Product",
        "Style",
        "Planned",
        "Completed",
        "Rejected",
        "Pending",
        "Stage",
        "Priority",
        "Start",
        "End",
        "Status",
    ]
    widths = [14, 22, 14, 10, 10, 10, 10, 12, 10, 12, 12, 12]
    row = _write_headers(ws, row, headers, widths)
    data_start = row

    total_planned = 0.0
    total_completed = 0.0
    total_rejected = 0.0

    for idx, rec in enumerate(records):
        planned = _num(rec.get("planned_quantity"))
        completed = _num(rec.get("completed_quantity"))
        rejected = _num(rec.get("rejected_quantity"))
        pending = planned - completed - rejected

        total_planned += planned
        total_completed += completed
        total_rejected += rejected

        data_row = [
            rec.get("production_number", ""),
            rec.get("product_name", ""),
            rec.get("style_name", ""),
            planned,
            completed,
            rejected,
            pending,
            rec.get("current_stage", ""),
            rec.get("priority", ""),
            _fmt_date(rec.get("planned_start")),
            _fmt_date(rec.get("planned_end")),
            rec.get("status", ""),
        ]
        row = _write_data_row(
            ws, row, data_row, number_cols={4, 5, 6, 7}, alt=idx % 2 == 1
        )

    _auto_filter(ws, data_start, row - 1, len(headers))

    efficiency = (total_completed / total_planned * 100) if total_planned > 0 else 0

    _write_total_row(
        ws,
        row,
        [
            "",
            "",
            "TOTAL",
            total_planned,
            total_completed,
            total_rejected,
            total_planned - total_completed - total_rejected,
            "",
            "",
            "",
            "",
            f"Eff: {efficiency:.1f}%",
        ],
        number_cols={4, 5, 6, 7},
    )

    return _save_workbook(wb)
